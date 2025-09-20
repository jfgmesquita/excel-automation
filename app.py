import openpyxl as xl
from openpyxl.chart import BarChart, Reference
import os

def process_workbook(filename, discount_rate=0.1, output_filename=None):
    """
    Process Excel workbook by applying discount to prices and creating a chart.
    
    Args:
        filename (str): Path to input Excel file
        discount_rate (float): Discount rate to apply (default 0.1 for 10%)
        output_filename (str): Output filename (default: adds '_corrected' to input)
    
    Returns:
        str: Path to the saved output file
    """
    # Validate input file exists
    if not os.path.exists(filename):
        raise FileNotFoundError(f"Input file '{filename}' not found")
    
    # Set default output filename
    if output_filename is None:
        base, ext = os.path.splitext(filename)
        output_filename = f"{base}_corrected{ext}"
    
    try:
        # Load workbook
        wb = xl.load_workbook(filename)
        
        # Use first sheet if "Sheet1" doesn't exist
        if "Sheet1" in wb.sheetnames:
            sheet = wb["Sheet1"]
        else:
            sheet = wb.active
            print(f"Warning: Using sheet '{sheet.title}' as 'Sheet1' was not found")
        
        # Add header for corrected prices if it doesn't exist
        if sheet.cell(1, 4).value is None:
            sheet.cell(1, 4).value = "Corrected Price"
        
        # Process data rows
        processed_count = 0
        processed_rows = []  # Track which rows have valid data
        
        for row in range(2, sheet.max_row + 1):
            price_cell = sheet.cell(row, 3)
            
            # Skip rows with no data or non-numeric values
            if price_cell.value is None:
                continue
                
            try:
                # Calculate corrected price
                original_price = float(price_cell.value)
                corrected_price = original_price * (1 - discount_rate)
                
                # Set corrected price in column 4
                corrected_price_cell = sheet.cell(row, 4)
                corrected_price_cell.value = corrected_price
                
                processed_count += 1
                processed_rows.append(row)  # Track this row as processed
                
            except (ValueError, TypeError):
                print(f"Warning: Skipping row {row} - invalid price value: {price_cell.value}")
                continue
        
        if processed_count == 0:
            print("Warning: No valid price data found to process")
            return None
        
        # Create chart with data labels and title
        # Use the full range from first to last processed row to ensure no gaps
        first_processed_row = min(processed_rows)
        last_processed_row = max(processed_rows)
        
        values = Reference(sheet, min_row=first_processed_row, max_row=last_processed_row, min_col=4, max_col=4)
        categories = Reference(sheet, min_row=first_processed_row, max_row=last_processed_row, min_col=1, max_col=1)
        
        chart = BarChart()
        chart.add_data(values, titles_from_data=False)
        chart.set_categories(categories)
        chart.title = f"Corrected Prices ({discount_rate*100}% Discount Applied)"
        chart.y_axis.title = "Price"
        chart.x_axis.title = "Items"
        
        print(f"Chart created with {processed_count} data points (rows {first_processed_row} to {last_processed_row})")
        
        # Add chart to sheet
        sheet.add_chart(chart, 'F2')
        
        # Save workbook
        wb.save(output_filename)
        print(f"Successfully processed {processed_count} items and saved to '{output_filename}'")
        
        return output_filename
        
    except Exception as e:
        raise Exception(f"Error processing workbook: {str(e)}")

# Example usage
if __name__ == "__main__":
    try:
        result = process_workbook("transactions.xlsx")
    except Exception as e:
        print(f"Error: {e}")
